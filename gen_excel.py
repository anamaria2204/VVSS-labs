#!/usr/bin/env python3
import sys
import os

# Minimal Excel generation using built-in capabilities
output_dir = '/Users/constantinmierla/Informatica/UBB-COMPUTER-SCIENCE/SEMESTRUL 6/Verificarea si validarea sistemelor soft/VVSS-labs/docs/Lab03'
output_file = os.path.join(output_dir, 'Lab03_WBT_TCs_Form.xlsx')

# Try multiple approaches
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    # Add test case data
    ws['A1'] = 'VVSS Lab03: White-Box Test Cases'
    ws['A2'] = 'StocService.areSuficient()'
    ws['A3'] = '2026-04-17'

    ws['A5'] = 'TC ID'
    ws['B5'] = 'Type'
    ws['C5'] = 'Input'
    ws['D5'] = 'Expected'
    ws['E5'] = 'Status'
    ws['F5'] = 'Coverage'

    test_data = [
        ['TC01', 'Valid', '3 ingredients sufficient', 'true', 'Pass', 'SC, DC, APC'],
        ['TC02', 'Invalid', 'First insufficient', 'false', 'Pass', 'DC, CC'],
        ['TC03', 'Invalid', 'Middle insufficient', 'false', 'Pass', 'LC, DCC'],
        ['TC04', 'Invalid', 'Last insufficient', 'false', 'Pass', 'LC, DCC'],
        ['TC05', 'Edge', 'Empty recipe', 'true', 'Pass', 'LC'],
        ['TC06', 'Boundary', 'Single ingredient exact', 'true', 'Pass', 'SC, DC'],
        ['TC07', 'Valid', 'Multiple entries sufficient', 'true', 'Pass', 'MCC'],
        ['TC08', 'Invalid', 'Multiple entries insufficient', 'false', 'Pass', 'MCC, DC'],
        ['TC09', 'Valid', 'Case-insensitive match', 'true', 'Pass', 'DCC'],
        ['TC10', 'Valid', 'Complex 4 ingredients', 'true', 'Pass', 'SC, DC'],
        ['TC11', 'Valid', 'Many ingredients (10)', 'true', 'Pass', 'LC, APC'],
        ['TC12', 'Invalid', 'Missing ingredient', 'false', 'Pass', 'DC, SC'],
        ['TC13', 'Boundary', 'Small quantities', 'true', 'Pass', 'SC, DC'],
        ['TC14', 'Boundary', 'Large quantities', 'true', 'Pass', 'SC, DC'],
        ['TC15', 'Valid', 'Statement coverage', 'true', 'Pass', 'SC'],
        ['TC16', 'Edge', 'Null recipe', 'Exception', 'Pass', 'SC'],
    ]

    for row_idx, row_data in enumerate(test_data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Add statistics sheet
    ws2 = wb.create_sheet("Statistics")
    ws2['A1'] = 'Test Execution Statistics'
    ws2['A3'] = 'Total Tests'
    ws2['B3'] = 16
    ws2['A4'] = 'Passed'
    ws2['B4'] = 16
    ws2['A5'] = 'Failed'
    ws2['B5'] = 0
    ws2['A6'] = 'Code Coverage %'
    ws2['B6'] = 100

    wb.save(output_file)
    print(f"✓ Created: {output_file}")
    sys.exit(0)

except ImportError:
    print("openpyxl not available, creating alternatives...")
    sys.exit(1)

