import sys
sys.path.insert(0, '/Users/constantinmierla/Informatica/UBB-COMPUTER-SCIENCE/SEMESTRUL 6/Verificarea si validarea sistemelor soft/VVSS-labs')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ==================== WORKSHEET 1: TITLE PAGE ====================
    title_sheet = wb.create_sheet("Title", 0)
    title_sheet.column_dimensions['A'].width = 20
    title_sheet.column_dimensions['B'].width = 50

    title_font = Font(name='Calibri', size=14, bold=True)
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font_white = Font(name='Calibri', size=14, bold=True, color="FFFFFF")

    # Header
    title_sheet['A1'] = 'VVSS Lab03'
    title_sheet['A1'].font = title_font_white
    title_sheet['A1'].fill = title_fill
    title_sheet.merge_cells('A1:B1')
    title_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    title_sheet['A3'] = 'Tema:'
    title_sheet['B3'] = 'Testare White-Box, Test Coverage'
    title_sheet['A4'] = 'Grupa:'
    title_sheet['B4'] = 'XXXX'
    title_sheet['A5'] = 'Membri echipa:'
    title_sheet['B5'] = 'Student 1, Student 2, Student 3'
    title_sheet['A6'] = 'Data:'
    title_sheet['B6'] = '2026-04-17'

    title_sheet['A8'] = 'Metoda testata:'
    title_sheet['B8'] = 'StocService.areSuficient(Reteta reteta)'

    title_sheet['A9'] = 'Functionalitate F02:'
    title_sheet['B9'] = 'Verificare stoc disponibil pentru ingrediente reteta'

    # Save
    wb.save('/Users/constantinmierla/Informatica/UBB-COMPUTER-SCIENCE/SEMESTRUL 6/Verificarea si validarea sistemelor soft/VVSS-labs/docs/Lab03/Lab03_WBT_TCs_Form.xlsx')
    print("Excel file created successfully!")
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()

